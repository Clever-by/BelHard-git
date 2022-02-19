# *** EXCEL

from openpyxl import load_workbook

wb = load_workbook(r"./taskWrk/resource/excelFile.xlsx")
print(f"sheetnames: {wb.sheetnames}")
sheet = wb['csvOut']
print(sheet.title)
print(sheet['A1'].value)
c = sheet['B2']
print(c.row, c.column, c.coordinate)

b = sheet.cell(row=2, column=2).value
print(b)

for i in range(1,4):
    print(i, sheet.cell(row=i, column=2).value)
    
    
# End