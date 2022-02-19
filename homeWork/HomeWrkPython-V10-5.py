# Сериализовать объект простого класса, загрузить его и запустить метод, использующий аттрибут объекта, при помощи pickle
# Создать таблицу Excel и конвертировать её в CSV при помощи Python.

from openpyxl import load_workbook
from random import randint
import csv
import json
import ast

wb = load_workbook(r"./homeWork/resource/excelFile.xlsx")
#print(f"sheetnames: {wb.sheetnames}")
sheet = wb['csvOut']
#print(sheet.title)
col_header = list()
i = 0
j = 0

for x in range(1,1048576):
    if sheet.cell(row=x, column=1).value is None:
        break
    value_row = sheet.cell(row=x, column=1).value
    #print(x, value_row)
    i += 1
print(f"row: {i}")

for y in range(1,16384):
    if sheet.cell(row=1, column=y).value is None:
        break
    value_column = sheet.cell(row=1, column=y).value
    #print(y, value_column)
    col_header.append(value_column)
    j += 1
print(f"column: {j}")
print(col_header)

#variant = randint(1,2)
variant = 2

if variant == 1: 
    x = 0
    string = ''
    spisok = list()
    while i >= x:
        x += 1
        y = 0
        #print(f"x={x}")
        if string != '':
            spisok.append(string[0:len(string)-1])
            string = ''
             
        while j >= y:
            y += 1
            #print(f"y={y}")
            string += str(sheet.cell(row=x, column=y).value) + ';'

        #print(string)

    print(spisok)
    
    print(f"Вариант обработки: {variant}")
    with open(r"./homeWork/resource/csvFile.csv", "w", encoding="utf-8") as wFile:
        for list_str in spisok:
            wFile.write(list_str+'\n')

    #wFile.close()

if variant == 2:
    d = dict()
    spisok = list()
    
    for x in range(2,i+1):
        for y in range(1,j+1):
            string = sheet.cell(row=x, column=y).value
            #print(col_header[(y-1)], string)
            d[col_header[y-1]]=string
            print(d)
            #print(len(d))
        if len(d) > 0:
            spisok.append(str(d))
            #print(f"spisok из d: {spisok}")
            d.clear()    
            
    
    #print(d)
    print(spisok)
    
    print(f"Вариант обработки: {variant}")        
    """
    # Параметр delimiter - указывает, какой использовать разделитель
    with open(r"./homeWork/resource/csvFileV2.csv", "w", encoding="utf-8") as wfile:
        writer = csv.DictWriter(wfile, fieldnames=list(spisok[0].keys()), delimiter=';')
        writer.writeheader()
        writer.writerows(spisok)
    """
    
    for str_dict in spisok:
        #print(str_dict)
        #print(json.loads(str_dict))
        #res_dict = json.loads(str_dict)
        print(ast.literal_eval(str_dict))
        #res_dict = json.loads(str_dict)
     

#End